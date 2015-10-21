#!/usr/bin/python

# Coding project for Mevoked
# Kaylee Moser
# Python version 2.7.10

# libraries
from __future__ import division
import sys
from openpyxl import load_workbook
import nltk
import random

def violence_features(tweet):
    features = {}
    # open bag of violent words
    wb1 = load_workbook('violent_bag.xlsx')
    ws_violence = wb1.active
    wb2 = load_workbook('swears.xlsx')
    ws_swears = wb2.active

    # definitions of stuff to return
    violent_words = []
    at_mentions = []
    swears = []
    exclamations = 0
    caps_words = 0

    # go through each word in tweet and check for different properties
    for word in tweet:
        found = 0
        # search for number of !
        exclamations += word.count('!')
        if (word == word.upper()):
           if (word is not 'I'):
              caps_words += 1 
        # if word is a mention, add to list
        if (word[0] == '@'):
           at_mentions.append(word) 
        else:
            # remove hashtags and punctuation, change to lower-case
            word = word.lower().strip(' ,#,,,.,!,?')
            for row in ws_violence.iter_rows('A2:A505'):
                if (word == row[0].value):
                    violent_words.append(word)
                    found = 1
                    break
            if (found == 0):
                for row in ws_swears.iter_rows('A2:A90'):
                    if (word == row[0].value):
                        swears.append(word)
                        break

    # add features to dictionary
    features['violent_freq'] = len(violent_words)
    #features['at_mentions'] = at_mentions
    features['exclamations'] = exclamations
    features['caps_percentage'] = caps_words/(len(tweet)-len(at_mentions))
    features['swears'] = len(swears)

    return features

def main():
    f1 = open('tweet_violent.txt', 'r')
    f2 = open('tweet_peaceful.txt', 'r')
    labeled_tweets = ([(tweet.split(), 'violent') for tweet in f1] + 
                      [(tweet.split(), 'peaceful') for tweet in f2])
    random.shuffle(labeled_tweets)
    
    featuresets = [(violence_features(tweet), VorP) for (tweet, VorP) in labeled_tweets]
    train_set, test_set = [featuresets[109:], featuresets[:109]]
    classifier = nltk.NaiveBayesClassifier.train(train_set)
    print "okay you need to fucking die"
    print classifier.classify(violence_features("okay you need to fucking die"))
    print classifier.show_most_informative_features(5)
    
    #tweet = sys.stdin.readline().split()
    #features = violence_features(tweet)
    #for key in features:
    #    print "{} : {}".format(key,features[key])

if (__name__ == '__main__'):
    main()
