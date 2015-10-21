#!/usr/bin/python

# Coding project for Mevoked
# Kaylee Moser
# Python version 2.7.10

# libraries
from __future__ import division
import sys
from openpyxl import load_workbook

def violence_features(tweet):
    features = {}
    # open bag of violent words
    wb = load_workbook('violent_bag.xlsx')
    ws = wb.active

    # definitions of stuff to return
    violent_words = []
    at_mentions = []
    exclamations = 0
    caps_words = 0

    # go through each word in tweet and check for different properties
    for word in tweet:
        # search for number of !
        exclamations += word.count('!')
        if (word == word.upper()):
           if (word is not 'I'):
              caps_words += 1 
        # if word is a mention, add to list
        if (word[0] == '@'):
           at_mentions.append(word) 
        else:
            # remove hashtags and punctuation, change to lower-case
            word = word.lower().strip(' ,#,,,.,!,?')
            for row in ws.iter_rows('A2:B495'):
                if (word == row[0].value):
                    violent_words.append(word)
                    break

    # add features to dictionary
    features['violent_words'] = violent_words
    features['violent_freq'] = len(violent_words)
    features['at_mentions'] = at_mentions
    features['exclamations'] = exclamations
    features['caps_percentage'] = caps_words/(len(tweet)-len(at_mentions))

    return features

def main():
    tweet = sys.stdin.readline().split()
    features = violence_features(tweet)
    for key in features:
        print "{} : {}".format(key,features[key])

if (__name__ == '__main__'):
    main()
