#!/usr/bin/python

# Coding project for Mevoked
# Kaylee Moser
# Python version 2.7.10

import sys
from openpyxl import load_workbook

# read in a sentence (tweet)
tweet = sys.stdin.readline().split()

# read in spreadsheet with values for pos/neg/neut
wb = load_workbook('test.xlsx')
ws = wb.active

#create lists for positive words and negative workds
pos = []
neg = []

# for each word in tweet, search spreadsheet for matching word
# NEEDS CHANGING; for now prints matching word
for word in tweet:
    for row in ws.iter_rows('A2:D20'):
        if (word == row[0].value):
            if (row[1].value == 1):
                print "negative: {}".format(word)
                neg.append(word)
            if (row[2].value == 1):
                print "positive: {}".format(word)
                pos.append(word)

print "negative = {}".format(neg)
print "positive = {}".format(pos)
                
                                
               


