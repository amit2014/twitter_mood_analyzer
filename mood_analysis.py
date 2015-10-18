#!/usr/bin/python

# Coding project for Mevoked
# Kaylee Moser
# Python version 2.7.10

# libraries
import sys
import tweepy
import string
import matplotlib.pyplot as plt
from openpyxl import load_workbook


def get_mood(tweet,ws):
    #create variables for positive words and negative words
    pos = 0
    neg = 0
    opposite = 0
    
    # for each word in tweet, search spreadsheet for matching word
    # if word matches, get negative/positive value and add to variable
    for word in tweet:
        # if word is a mention, skip
        if (word[0] == '@'):
            opposite = 0
        else:
            # remove hashtags and punctuation, change to lower-case
            word = word.lower().strip(' ,#,,,.,!,?')
            if (word == 'not'):
                opposite = 1
            else:
                for row in ws.iter_rows('A1:B253'):
                    if (word == row[0].value):
                        if (row[1].value == 'negative'):
                            if (opposite == 1):
                                pos += 1
                                opposite = 0
                            else:
                                neg += 1
                        else:
                            if (opposite == 1):
                                neg += 1
                                opposite = 0
                            else:
                                pos +=1
                        break
    
    # find mood number by taking total number of positive words and subtracting total
    # number of negative words
    mood_number = pos - neg
    return mood_number

def plot_mood(dates,moods):
    plt.plot(dates,moods)
    plt.ylabel('Mood Level')
    plt.show()


def main():
    # consumer keys and access tokens, used for OAuth
    consumer_key = 'ArJTlGOeA7rnIHyiPZkHmMhIJ'
    consumer_secret = 'zr7Ldopc0QdwxvsMwBcFFYk3LsWUPeuKEkfo6wj9848W4x7eq0'
    access_token = '1550061679-WpJunbZbtwCGXL84wQbyzmOuSre1tNh7iloJyQV'
    access_token_secret = 'PusPCv8NgalPKSagKNMioU8xpP77TrhSXS9wKH0A5K8bE'
    
    # OAuth process, using the keys and tokens
    auth = tweepy.OAuthHandler(consumer_key, consumer_secret)
    auth.set_access_token(access_token, access_token_secret)
    
    # Creation of actual interface, using authentication
    api = tweepy.API(auth)

    # Get Twitter user name from program user
    screen_name = raw_input("Please enter the Twitter handle: ")
    
    # Open spreadsheet of values
    wb = load_workbook('bag_of_words.xlsx')
    ws = wb.active

    # Gather recent tweets, split each tweet into a list of words
    recent_tweets = api.user_timeline(screen_name)
    moods = []
    dates = []
    for tweet in recent_tweets:
        moods.append(get_mood(tweet.text.split(),ws))
        dates.append(tweet.created_at)

    for mood in range(0,len(moods)):
        if (moods[mood] <= -3):
            print "ALERT! Negative tweet:"
            print recent_tweets[mood].created_at
            print recent_tweets[mood].text
    print sum(moods)
    if (sum(moods) < 0):
        print "ALERT! Tweets are generally negative"

    plot_mood(dates,moods)


if (__name__ == '__main__'):
    main()


