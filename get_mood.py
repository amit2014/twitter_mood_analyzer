#!/usr/bin/python

# Coding project for Mevoked
# Kaylee Moser
# Python version 2.7.10

import sys
from openpyxl import load_workbook

# read in a sentence (tweet)
tweet = sys.stdin.readline().split()

# read in spreadsheet with values for pos/neg/neut
wb = load_workbook('test.xlsx')
ws = wb.active

#create lists for positive words and negative workds
pos = 0
neg = 0

# for each word in tweet, search spreadsheet for matching word
# if word matches, get negative/positive value and add to variable
for word in tweet:
    for row in ws.iter_rows('A2:D20'):
        if (word == row[0].value):
            if (row[1].value == 1):
                neg += 1
            if (row[2].value == 1):
                pos +=1

# find mood number by taking total number of positive words and subtracting total
# number of negative words
mood_number = pos - neg
print mood_number

                
                                
               


